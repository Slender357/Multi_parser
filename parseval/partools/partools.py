import csv
import openpyxl
import httplib2
import time
import apiclient
from datetime import datetime
from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from oauth2client.service_account import ServiceAccountCredentials
from googleapiclient.http import MediaIoBaseDownload
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
import io
import yaml
import os
import logging
import chromedriver_autoinstaller
from telebot import TeleBot
from .config import CONFIG
from functools import wraps


def decorator(func):
    @wraps(func)
    def inner(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except BaseException as r:
            save_to_log('Ошибка фукции ' + func.__name__ + ': ' + str(r), 'error')

    return inner


def start_end(func):
    @wraps(func)
    def inner(*args, **kwargs):
        save_to_log(func.__name__ + ' start', 'info')
        m = func(*args, **kwargs)
        if m:
            save_to_log(func.__name__ + ' ok', 'info')
        else:
            save_to_log('Запись данных ' + func.__name__ + ' не произведена', 'warning')
        return m

    return inner


@decorator
def sent_to_tbot(dic, telegram_bot_token=CONFIG['telegram_bot_token'], chat_id=CONFIG['chat_id']):
    bot = TeleBot(telegram_bot_token)
    bot.config['api_key'] = telegram_bot_token
    text = datetime.today().strftime("%Y-%m-%d %H:%M:%S") + '\n'
    for i in dic:
        value = list(i.values())[0]
        key = list(i.keys())[0]
        if value:
            text += key + ': ОК' + '\n'
        else:
            text += key + ': ERROR' + '\n'
    bot.send_message(chat_id, text)
    try:
        with open('telegram.txt', 'r') as f:
            massenge = set()
            for i in f.readlines():
                massenge.add(i)
        text = 'При загрузке данных были добавлены комментарии:' + '\n'
        for i in massenge:
            text += i
        bot.send_message(chat_id, text)
        os.remove('./telegram.txt')
    except FileNotFoundError:
        bot.send_message(chat_id, 'Все данные корректны')


@decorator
def er_for_telegramm(er, shet):
    with open('telegram.txt', 'a') as f:
        f.write(shet + ' : ' + er + '\n')
    pass


@decorator
def save_to_log(text, metod):
    logging.basicConfig(format=u'%(levelname)-8s [%(asctime)s] %(message)s', level=logging.DEBUG, filename=u'log.log')
    logging.disable(logging.DEBUG)

    if metod == 'info':
        logging.info(text)
    if metod == 'warning':
        logging.warning(text)
    if metod == 'error':
        logging.error(text)


@decorator
def quickstart_sheet(spreadsheet_id=CONFIG['spreadsheet_id'], credentials_file=CONFIG['credentials_file'],
                     type_conection='sheets',
                     version_conection='v4'):
    credentials = ServiceAccountCredentials.from_json_keyfile_name(
        credentials_file,
        ['https://www.googleapis.com/auth/spreadsheets',
         'https://www.googleapis.com/auth/drive'])
    httpAuth = credentials.authorize(httplib2.Http())
    service = apiclient.discovery.build(type_conection, version_conection, http=httpAuth)
    return spreadsheet_id, service


@decorator
def update_wialon_token(new_token, CONFIG):
    CONFIG.update({'token_wialon': new_token})
    with open("config.yaml", "w", encoding='utf8') as f:
        yaml.dump(CONFIG, f)


@decorator
def get_days(listsheet):
    spreadsheet_id, service = quickstart_sheet()
    last_dates = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=listsheet + '!A1:A9999999'

    ).execute()
    last_date = datetime.strptime(last_dates['values'][len(last_dates['values']) - 1][0], '%d.%m.%Y')
    today_date = datetime.strptime(datetime.strftime(datetime.today(), '%d.%m.%Y'), '%d.%m.%Y')
    delta_day = int((today_date - last_date).days - 1)
    return delta_day


class Login:
    def __init__(self, param):
        if param == 'Autodor':
            self.f_site_url = CONFIG['autodor_first_site_url']
            self.username = CONFIG['autodor_username']
            self.password = CONFIG['autodor_password']
            self.usn_xpath = CONFIG['autodor_username_xpath']
            self.pw_xpath = CONFIG['autodor_password_xpath']
            self.bt_xpath = CONFIG['autodor_login_button_xpath']
            self.second_site_url = CONFIG['autodor_second_site_url']
            self.second_click = None
        if param == 'Platon':
            self.f_site_url = CONFIG['platon_first_site_url']
            self.username = CONFIG['platon_username']
            self.password = CONFIG['platon_password']
            self.usn_xpath = CONFIG['platon_username_xpath']
            self.pw_xpath = CONFIG['platon_password_xpath']
            self.bt_xpath = CONFIG['platon_login_button_xpath']
            self.second_site_url = CONFIG['platon_second_site_url']
            self.second_click = CONFIG['platon_clik_to_serch_xpath']
        if param == 'Wialon':
            self.f_site_url = CONFIG['wialon_first_site_url']
            self.username = CONFIG['wialon_username']
            self.password = CONFIG['wialon_password']
            self.usn_xpath = CONFIG['wialon_username_xpath']
            self.pw_xpath = CONFIG['wialon_password_xpath']
            self.bt_xpath = CONFIG['wialon_login_button_xpath']
            self.second_site_url = None
            self.second_click = None

    @decorator
    def login_to_site(self):
        caps = DesiredCapabilities().CHROME
        caps["pageLoadStrategy"] = "eager"
        options = webdriver.ChromeOptions()
        options.add_argument('headless')
        options.add_argument('window-size=1920x935')
        driver = webdriver.Chrome(desired_capabilities=caps, options=options,
                                  executable_path=chromedriver_autoinstaller.install())
        driver.get(self.f_site_url)
        time.sleep(5)
        driver.find_element(By.XPATH, self.usn_xpath).send_keys(self.username)
        driver.find_element(By.XPATH, self.pw_xpath).send_keys(self.password)
        time.sleep(1)
        driver.find_element(By.XPATH, self.bt_xpath).click()
        time.sleep(3)
        if self.second_site_url is not None:
            driver.get(self.second_site_url)
            time.sleep(5)
        if self.second_click is not None:
            driver.find_element(By.XPATH, self.second_click).click()
            time.sleep(5)
        return driver


@decorator
def get_sheet_values(sheet):
    spreadsheet_id, service = quickstart_sheet()
    sheet_values = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=sheet + '!A1:Z9999999'
    ).execute()
    keys = sheet_values['values'][0]
    dic_values = []
    for k in sheet_values['values']:
        d = {}
        m = 0
        for i in keys:
            try:
                d.update({i: k[m]})
                m += 1
            except BaseException:
                d.update({i: ''})
                m += 1
        dic_values.append(d)
    dic_values.pop(0)
    return dic_values


@decorator
def reform_date(tr3):
    tr3 = datetime.strptime(int_value_from_ru_month(tr3), "       %d %m %H:%M      ")
    if tr3.month <= datetime.today().month:
        tr3 = datetime.strptime(datetime.strftime(tr3,
                                                  '%d.%m') + '.' + str(
            datetime.today().year), '%d.%m.%Y')
    else:
        tr3 = datetime.strptime(datetime.strftime(tr3,
                                                  '%d.%m') + '.' + str(
            datetime.today().year - 1), '%d.%m.%Y')
    return tr3


@decorator
def get_tclist(tcsheet):
    spreadsheet_id, service = quickstart_sheet()
    tclist = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=tcsheet + '!A1:C9999999'
    ).execute()
    own_park = {'ТС': []}
    renters = {'ТС': []}
    hired_car = {'ТС': [], 'Перевозчик': []}
    for i in tclist['values']:
        if i[1] == 'own_park':
            own_park['ТС'].append(i[0])
        elif i[1] == 'renters':
            renters['ТС'].append(i[0])
        elif i[1] == 'hired_car':
            hired_car['ТС'].append(i[0])
            try:
                hired_car['Перевозчик'].append(i[2])
            except IndexError:
                hired_car['Перевозчик'].append('')
    return own_park, renters, hired_car


@decorator
def int_value_from_ru_month(date_str):
    RU_MONTH_VALUES = {
        'января': '01',
        'янв.': '01',
        'февраля': '02',
        'февр.': '02',
        'марта': '03',
        'апреля': '04',
        'апр.': '04',
        'мая': '05',
        'июня': '06',
        'июля': '07',
        'августа': '08',
        'авг.': '08',
        'сентября': '09',
        'сент.': '09',
        'октября': '10',
        'окт.': '10',
        'ноября': '11',
        'нояб.': '11',
        'декабря': '12',
        'дек.': '12'
    }
    for k, v in RU_MONTH_VALUES.items():
        date_str = date_str.replace(k, str(v))
    return date_str


@decorator
def chek_date(date, order):
    date = datetime.strptime(int_value_from_ru_month(date), "%d %m")
    if datetime.today().month == 11 or datetime.today().month == 12:
        if date.month == 1 or date.month == 2:
            year = datetime.today().year
            s2 = (datetime.strftime(
                date, '%d.%m') + '.' + str(
                year + 1))
        else:
            year = datetime.today().year
            s2 = (datetime.strftime(
                date, '%d.%m') + '.' + str(
                year))
    elif datetime.today().month == 1 or datetime.today().month == 2:
        if date.month == 11 or date.month == 12:
            year = datetime.today().year - 1
            s2 = (datetime.strftime(
                date, '%d.%m') + '.' + str(
                year))
        elif order > 10000:
            year = datetime.today().year - 1
            s2 = (datetime.strftime(
                date, '%d.%m') + '.' + str(
                year + 1))
        else:
            year = datetime.today().year
            s2 = (datetime.strftime(
                date, '%d.%m') + '.' + str(
                year))

    else:
        year = datetime.today().year
        s2 = (datetime.strftime(
            date, '%d.%m') + '.' + str(
            year))

    return year, s2


@decorator
def razdelenie(proverca_dat_list, dic, listsheet):
    body = {'valueInputOption': 'RAW',
            'data': []
            }
    for k in proverca_dat_list:
        for m in dic:
            if m['Заявка'] == int(list(k.values())[0]) and m['Год заявки'] == int(list(k.values())[1]):
                body['data'].append({
                    'range': listsheet + '!' + list(k.keys())[0],
                    'values': [list(m.values())]
                })
                dic.remove(m)
    return dic, body


@decorator
def safe_file_csv(dic, name):
    with open(name + '.csv', 'w', newline='') as file:
        writer = csv.writer(file, delimiter=';')
        # writer.writerow(product[0].keys())
        for item in dic:
            writer.writerow(item.values())


@decorator
def safe_file_xls(dic, name, sheet, tdate):
    wb = openpyxl.load_workbook(name + '.xlsx')
    ws = wb[sheet]
    k = 0
    for i in range(1, ws.max_row + 1):
        if ws.cell(row=i, column=1).value == tdate:
            k += 1
    if k >= 1:
        print('В листе ' + sheet + ' есть дата ' + tdate + ' ,количество строк - ' + str(k))
    else:
        for i in dic:
            data = list(i.values())
            ws.append(data)
        wb.save(name + '.xlsx')


@decorator
def save_to_sheetlist(dic, service, spreadsheet_id, listsheet):
    d = []
    for i in dic:
        k = []
        for o in i:
            k.append(i[o])
        d.append(k)
    service.spreadsheets().values().append(
        spreadsheetId=spreadsheet_id,
        range=listsheet + '!A1',
        valueInputOption='RAW',
        body={'values': d
              }
    ).execute()
    return True


@decorator
def safe_to_sheet(dic, listsheet, proverca=0):
    spreadsheet_id, service = quickstart_sheet()
    if proverca == 1:
        if dic == []:
            return False
        proverca_dat_dict = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=listsheet + '!A1:A9999999'

        ).execute()
        proverca_dat_set = set()
        for i in proverca_dat_dict['values']:
            proverca_dat_set.add(i[0])
        num = 0
        for k in dic:
            if k['Дата'] in proverca_dat_set:
                num += 1
        if num >= 1:
            print('В листе ' + listsheet + ' уже есть добавляемые даты в количестве ' + str(num))
            return False
        else:
            return save_to_sheetlist(dic, service, spreadsheet_id, listsheet)
    if proverca == 2:
        proverca_dat_dict = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=listsheet + '!A1:B9999999'

        ).execute()
        proverca_dat_list = []
        schet = 1
        for i in proverca_dat_dict['values']:
            proverca_dat_list.append({
                'A' + str(schet): i[0],
                'B' + str(schet): i[1]
            })
            schet += 1
        proverca_dat_list.pop(0)
        dic, body = razdelenie(proverca_dat_list, dic, listsheet)
        d = []
        for i in dic:
            k = []
            for o in i:
                k.append(i[o])
            d.append(k)
        if d == []:
            return False
        else:
            service.spreadsheets().values().batchUpdate(spreadsheetId=spreadsheet_id, body=body).execute()
            service.spreadsheets().values().append(
                spreadsheetId=spreadsheet_id,
                range=listsheet + '!A1',
                valueInputOption='RAW',
                body={'values': d
                      }
            ).execute()
            os.remove(CONFIG['file_1c'])
            return True


@decorator
def save_file_for_1c(path_to_save):
    spreadsheet_id, service = quickstart_sheet(type_conection='drive',
                                               version_conection='v3')
    spisochek = service.files().list(pageSize=10,
                                     fields="nextPageToken, files(id, name, mimeType)").execute()
    file_id = None
    for i in spisochek['files']:
        if CONFIG['file_1c'] == i['name']:
            file_id = i['id']
            break
    request = service.files().get_media(fileId=file_id)
    filename = path_to_save
    fh = io.FileIO(filename, 'wb')
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
    return
